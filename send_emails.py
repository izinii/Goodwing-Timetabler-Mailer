import smtplib
import pandas as pd
import os
import unicodedata
from email.message import EmailMessage
from openpyxl import load_workbook


# Load the original Excel file
file_path = "path_to_the_file/name_of_the_file.xlsx" # Replace with the path to your Excel file
xls = pd.ExcelFile(file_path)

# Find all professor sheets
teacher_sheets = [sheet for sheet in xls.sheet_names if sheet.startswith("Teacher")]

# Function to normalize names (remove accents)
def normalize_name(name):
    return ''.join(c for c in unicodedata.normalize('NFD', name) if unicodedata.category(c) != 'Mn')

# Retrieve professors' emails
professors = {}
for sheet in teacher_sheets:
    parts = sheet.split("_")
    last_name = normalize_name(parts[1].lower())   # remove accents and lowercase
    first_name = normalize_name(parts[2].lower())  # remove accents and lowercase
    email = f"{first_name}.{last_name}@edu.devinci.fr"

    if email not in professors:
        professors[email] = []

    professors[email].append(sheet)

# Email configuration
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_SENDER = "your.email@gmail.com"  # Replace with your email
EMAIL_PASSWORD = "your_app_password"  # Replace with your generated app password


def send_email(recipient, schedules):
    """
    Sends an email to the professor with their timetable attached.
    """
    msg = EmailMessage()
    msg["Subject"] = "Your University Timetable"
    msg["From"] = EMAIL_SENDER
    msg["To"] = recipient
    
    last_name_teacher = recipient.split("@")[0].split(".")[1] # Extract last name of an email address (second part before '@')
    msg.set_content(
        f"Dear Professor {last_name_teacher.capitalize()},\n\n"
        f"Please find attached your timetable for the next semester.\n\n"
        f"Best regards,\n"
        f"University Administration"
    )

    # Create a TEMPORARY Excel file with only the professor's sheets
    temp_file = f"temp_schedule_{recipient}.xlsx"
    wb_new = load_workbook(file_path)  # Load the original Excel file

    # Remove unnecessary sheets
    for sheet in wb_new.sheetnames:
        if sheet not in schedules:
            wb_new.remove(wb_new[sheet])

    wb_new.save(temp_file)  # Save the new Excel file with only the necessary sheets

    # Attach the formatted timetable file
    with open(temp_file, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="timetable.xlsx"
        )

    # Send the email
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
            print(f"✅ Email sent to {recipient}")
    except Exception as e:
        print(f"❌ Error sending to {recipient}: {e}")

    # Delete the temporary file
    os.remove(temp_file)


# Send emails to all professors
for email, sheets in professors.items():
    send_email(email, sheets)

print("🎉 All emails have been sent successfully!")